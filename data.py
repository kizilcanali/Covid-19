import requests
from openpyxl import Workbook, load_workbook
import datetime


"""
    200: Everything went okay, and the result has been returned (if any).
    301: The server is redirecting you to a different endpoint.
    This can happen when a company switches domain names,
    or an endpoint name is changed.
    400: The server thinks you made a bad request.
    This can happen when you don’t send along the right data,
    among other things.
    401: The server thinks you’re not authenticated.
    Many APIs require login ccredentials,
    so this happens when you don’t send the right credentials to access an API.
    403: The resource you’re trying to access is forbidden:
    you don’t have the right permissions to see it.
    404: The resource you tried to access wasn’t found on the server.
    503: The server is not ready to handle the request.
"""


def create_data_excel(name):
    """
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Date"
    sheet["B1"] = "Country"
    sheet["C1"] = "Confirmed"
    sheet["D1"] = "Deaths"
    sheet["E1"] = "Recovered"

    workbook.save(filename="{}.xlsx".format(name))


def edit_date(unedited_date):
    """
    A function used to edit date coming from API

    Parameters
    ----------
    ``unedited_date`` : ``str``
        The date of the data

    Return
    ------
    ``Year-Month-Day`` -> ``2020-02-03``
    """
    edited_date = unedited_date
    date = edited_date.split("T")
    return date[0]


def total_status_by_country(country):
    datas = []
    api = requests.get(
        'https://api.covid19api.com/live/country/{country}/status/confirmed'.format(country=country))
    if api.json() != []:
        data = api.json()[-1]
        datas.append(data["Confirmed"])
        datas.append(data["Deaths"])
        datas.append(data["Recovered"])
        return datas


def total_status_by_country_excel(country, excel_name):
    """
    A function shows the confirmed, deaths, recovered,
    active datas of ``country``

    Parameters
    ----------
    ``country`` : ``str``
        The country whose data are desired to be seen

    ``excel_name``: ``str``
        Excel file name

    Return
    ------
    ``status{'Confirmed': confirmed_data, 'Deaths': deaths_data,
    'Recovered': recovered_data, 'Active': active_data}``
    """
    api = requests.get(
        'https://api.covid19api.com/live/country/{country}/status/confirmed'.format(country=country))
    if api.json() != []:
        data = api.json()[-1]

        create_data_excel(excel_name)

        workbook = load_workbook('{}.xlsx'.format(excel_name))
        sheet = workbook.active
        sheet["A2"] = edit_date(data['Date'])
        sheet["B2"] = country
        sheet["C2"] = data['Confirmed']
        sheet["D2"] = data['Deaths']
        sheet["E2"] = data['Recovered']
        workbook.save(filename='{}.xlsx'.format(excel_name))
        workbook.close()
        return True
    else:
        return False


def just_global_excel():
    api = requests.get('https://api.covid19api.com/summary')
    data = api.json()['Global']
    create_data_excel("global")

    workbook = load_workbook("global.xlsx")
    sheet = workbook.active
    sheet["A2"] = datetime.date.today()
    sheet["B2"] = "Global"
    sheet["C2"] = data['TotalConfirmed']
    sheet["D2"] = data['NewDeaths']
    sheet["E2"] = data['TotalRecovered']

    workbook.save(filename="global.xlsx")
    workbook.close()


def just_global():
    datas = []
    api = requests.get('https://api.covid19api.com/summary')
    data = api.json()['Global']
    datas.append(data['TotalConfirmed'])
    datas.append(data['NewDeaths'])
    datas.append(data['TotalRecovered'])
    return datas


def total_status_by_global():
    """
    """
    api = requests.get('https://api.covid19api.com/summary')
    data_countries = api.json()['Countries']

    column_counter = 3

    workbook = load_workbook("global.xlsx")
    sheet = workbook.active

    for d in data_countries:
        sheet["A{}".format(column_counter)] = datetime.date.today()
        sheet["B{}".format(column_counter)] = d['Country']
        sheet["C{}".format(column_counter)] = d['TotalConfirmed']
        sheet["D{}".format(column_counter)] = d['TotalDeaths']
        sheet["E{}".format(column_counter)] = d['TotalRecovered']
        column_counter += 1

    workbook.save(filename="global.xlsx")
    workbook.close()


def getCountries():
    api = requests.get('https://api.covid19api.com/summary')
    data_countries = api.json()['Countries']
    gcountries = []
    gcountries.append("Global")
    for i in data_countries:
        gcountries.append(i['Country'])

    return gcountries
